



from openpyxl import load_workbook
import pickle


dati = {}

ws = load_workbook('Acc_Rappresentativi_normalizzati.xlsx', data_only=True).active
n_colonne = ws.max_column
dati['acc'] = {}
for col_idx in range(1, n_colonne + 1):
    col_values = []
    for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        col_values.append(val)
    dati['acc'][f'{col_idx}'] = col_values


ws = load_workbook('Sel_Rappresentativi_normalizzati.xlsx', data_only=True).active
n_colonne = ws.max_column
dati['spectr'] = {}
# first column as T
col_values = []
for row in ws.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):
    val = row[0]
    col_values.append(val)
dati['spectr']["T"] = col_values


for col_idx in range(2, n_colonne + 1):
    col_values = []
    for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True):
        val = row[0]
        col_values.append(val)
    dati['spectr'][f'{col_idx-1}'] = col_values

with open('acc_data.pkl', 'wb') as f:
    pickle.dump(dati, f)
