




import openseespy.opensees as ops
import matplotlib.pyplot as plt
try:
    import plot_settings as plt_set
except ImportError:
    class plt_set:
        pass
    plt_set.fig_size = (6, 4.5) # default figure size if plot_settings is not available
import numpy as np
import pickle
from tqdm import tqdm
import openpyxl
import os




def os_analysis(id_acc, T_1, q, accs_data): 
    """Run OpenSees analysis for a given acceleration record and period.
    Args:
        id_acc (int): ID of the acceleration record.
        T_1 (float): Period for the analysis.
        q (float): Behavior factor.
        accs_data (dict): Dictionary containing acceleration data and spectral information (from load_acc_from_excel.py).
    Returns:
        float: Maximum displacement response for the given period and behavior factor.
    """

    global dy # da ripensare 

    # unit of measurement: N, mm, sec
    ops.wipe()
    ops.model('basic', '-ndm', 1, '-ndf', 1)

    # Nodes
    ops.node(1, 0), ops.fix(1, 1)
    ops.node(11, 0)


    # Material
    fy = 57382058.84  # Yield stress [N]
    fmaxfy=1.13
    fresfy=0
    k = 2193503.77  # Young's modulus [N/mm]
    dy=fy/k
    dp=dy*1.3
    # thetap= 60.2
    dpc= 183
    du=350


    ops.uniaxialMaterial ('IMKPinching', 2, k, dp, dpc, du, fy, fmaxfy, fresfy, -dp, -dpc, -du, -fy, fmaxfy, fresfy, 400, 1000, 700, 400, 1, 1, 1, 1, 1, 1, 0.5, 0.5)
    ops.uniaxialMaterial('MinMax', 1, 2, '-min', -2.5*du, '-max', 2.5*du) 


    # Element
    ops.element('zeroLength', 1, 1, 11, '-mat', 1, '-dir', 1)



    #
    #  NLTH analysis
    #
    g = 9806.65  # mm/s²

    # Massa
    Mnode = k*T_1**2/(4*np.pi**2)
    ops.mass(11, Mnode, 0, 0)

    # # Modal analysis (adesso è commentato perché non serve per il calcolo dello spettro)
    # eigenvalues = ops.eigen('-fullGenLapack',1)
    # periods = [(2*np.pi)/np.sqrt(lam) for lam in eigenvalues]
    # for i, freq in enumerate(periods):
    #     print(f"Modo {i+1}: frequenza = {freq:.4f} sec")

    # Leggere i dati di accelerazione
    acc_data = np.array(accs_data["acc"][f'{id_acc}']) * g #converting acceleration to mm/s²

    # take position of accs_data["spectr"]["T"] equal to T_1
    T_index = accs_data["spectr"]["T"].index(T_1)
    # Prendere pseudo accelerazione per T_1
    S_el = accs_data["spectr"][f'{id_acc}'][T_index]  # spectral acceleration for T_1

    # Scale factor for acceleration
    a_g = (k * dy * q) / (Mnode * S_el * g)

    # Damping
    xi = 5/100
    alphas = 2*xi*np.sqrt(k/Mnode)
    ops.rayleigh(alphas, 0, 0, 0)

    # Parameters for the analysis
    dt = 0.005
    step_dt = 1
    DtAnalysis = dt/step_dt
    TMaxAnalysis = len(acc_data) * dt
    dof_analysis = 1

    # Impostare parametri di analisi
    ops.test('EnergyIncr', 1e-4, 20, 0)
    ops.constraints('Transformation')
    ops.numberer('RCM')
    ops.system("BandGen")
    ops.algorithm('Newton')
    ops.integrator('Newmark', 0.5, 0.25)
    ops.analysis('Transient')
    ops.timeSeries('Path', 2, '-dt', dt, '-values', *acc_data, '-factor', a_g)
    ops.pattern('UniformExcitation', 10, dof_analysis, '-accel', 2)


    # Integrazione dell'equazione del moto
    tCurrent = ops.getTime()
    time, D, ok = [], [], 0

    while ok == 0 and tCurrent < TMaxAnalysis:
        ok = ops.analyze(1, DtAnalysis)

        if ok != 0:
            ops.test('NormDispIncr', 1e-4, 100, 0)
            ops.algorithm('ModifiedNewton', '-initial')
            ok = ops.analyze(1, DtAnalysis)
            if ok == 0:
                ops.test('EnergyIncr', 1e-4, 20, 0)
                ops.algorithm('Newton')
            # else:
            #     print(f"Analysis accel:{id_acc:i}, T:{T_1:.2f}, q:{q:.2f} failed!")
        
        tCurrent = ops.getTime()
        time.append(tCurrent)

        D.append(ops.nodeDisp(11, dof_analysis))
        
    return abs(max(D, key=abs))


def eval_spectrum(id_acc, q, accs_data, **kwargs):
    """Evaluate the displacement spectrum for a given acceleration record and scaling factor.
    Args:
        id_acc (int): ID of the acceleration record.
        q (float): Behavior factor for the spectral displacement.
        accs_data (dict): Dictionary containing acceleration data and spectral information (from load_acc_from_excel.py).
        **kwargs: Additional keyword arguments, e.g., Ts for periods.
    Returns:
        list: List of spectral displacements for the given periods.
    """

    Ts = kwargs.get('Ts', np.arange(0.10, 4, 0.05)) #  default periods if not provided
    return [os_analysis(id_acc, round(T, 2), q, accs_data)/dy for T in Ts]


def export_to_excel(S_ds, filename='displ_spectra_results.xlsx'):
    """Export the spectral displacement results to an Excel file.
    Args:
        S_ds (dict): Dictionary containing spectral displacements for different behavior factors.
        filename (str): Name of the output Excel file.
    """
    # Crea un nuovo file Excel
    wb = openpyxl.Workbook()

    # Per ogni accelerogramma (id_acc), crea un foglio
    for id_acc in range(1, len(S_ds[f"{qs[0]}"])+1):
        ws = wb.create_sheet(title=f"acc_{id_acc}")
        # Prima colonna: T
        ws.cell(row=1, column=1, value="T [s]")
        for i, T in enumerate(Ts, start=2):
            ws.cell(row=i, column=1, value=T)
        # Colonne successive: q e valori S_ds
        for col_idx, q in enumerate(qs, start=2):
            ws.cell(row=1, column=col_idx, value=f"q = {q}")
            for row_idx, val in enumerate(S_ds[f"{q}"][id_acc-1], start=2):
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Aggiungi un foglio per la media degli spettri
    ws_avg = wb.create_sheet(title="avg_spectra")
    # Prima colonna: T
    ws_avg.cell(row=1, column=1, value="T [s]")
    for i, T in enumerate(Ts, start=2):
        ws_avg.cell(row=i, column=1, value=T)
    # Colonne successive: q e valori medi
    for col_idx, q in enumerate(qs, start=2):
        ws_avg.cell(row=1, column=col_idx, value=f"q = {q}")
        for row_idx, val in enumerate(S_ds[f"{q}_avg"], start=2):
            ws_avg.cell(row=row_idx, column=col_idx, value=val)

    # Rimuovi il foglio di default creato da openpyxl se non usato
    if 'Sheet' in wb.sheetnames:
        std = wb['Sheet']
        wb.remove(std)

    # Salva il file
    wb.save(filename)
    pass


def plot_spectrum(fig, ax, S_ds, Ts, q):
    """Plot the spectral displacement for a given scaling factor.
    Args:
        fig (Figure): Matplotlib figure object.
        ax (Axes): Matplotlib axes object.
        S_ds (dict): Dictionary containing spectral displacements.
        Ts (list): List of periods.
        q (float): Behavior factor for the spectral displacement.
    """
    for S_d in S_ds[f"{q}"]: # plot di tutti gli spettri
        ax.plot(Ts, S_d, c="midnightblue", linewidth=0.5, alpha=0.4)
    # Plot of the average
    ax.plot(Ts, S_ds[f"{q}_avg"], c="midnightblue", linewidth=1.5, label="Avg spectr.")
    ax.set_xlabel('T [s]'), ax.set_ylabel(r'$S_d/d_y$ [-]')
    ax.grid(alpha = 0.3)
    ax.axhline(q, color = "dimgrey", linestyle='--', linewidth=1.5, label=rf'$\mu$ = {q}')
    ax.legend()
    ax.set_xlim([0.1, 4]), ax.set_ylim([0, q*1.75])
    return fig, ax


def load_acc_from_excel(filename='acc_data.pkl'):
    """Load acceleration data and spectral information from Excel files and save them to a pickle file.
    This function reads two Excel files: 'Acc_Rappresentativi_normalizzati.xlsx' and 'Sel_Rappresentativi_normalizzati.xlsx',
    extracts the relevant data, and saves it in a dictionary format to 'acc_data.pkl'.
    """

    dati = {}

    ws = openpyxl.load_workbook('Acc_Rappresentativi_normalizzati.xlsx', data_only=True).active
    n_colonne = ws.max_column
    dati['acc'] = {}
    for col_idx in range(1, n_colonne + 1):
        col_values = []
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            col_values.append(val) if val is not None else None
        dati['acc'][f'{col_idx}'] = col_values


    ws = openpyxl.load_workbook('Sel_Rappresentativi_normalizzati.xlsx', data_only=True).active
    n_colonne = ws.max_column
    dati['spectr'] = {}
    # first column as T
    col_values = []
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=1, values_only=True):
        val = round(row[0], 2) # round to 2 decimal places
        col_values.append(val)
    dati['spectr']["T"] = col_values


    for col_idx in range(2, n_colonne + 1):
        col_values = []
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx, values_only=True):
            val = row[0]
            col_values.append(val)
        dati['spectr'][f'{col_idx-1}'] = col_values

    with open(filename, 'wb') as f:
        pickle.dump(dati, f)
    pass

if __name__ == "__main__": # main function starts here

    n_gm = 7  # Numero di accelerogrammi

    # Carica i dati di accelerazione da Excel se non esiste il file pickle
    load_acc_from_excel() if not os.path.exists('acc_data.pkl') else None

    # Crea la cartella figs se non esiste
    os.makedirs('./figs', exist_ok=True)

    # read acceleration data from pickle file
    with open('acc_data.pkl', 'rb') as f:
        accs_data = pickle.load(f)
    
    # Range periodi su cui calcolare lo spettro
    Ts = np.arange(0.1, 4, 0.05)
    qs = np.arange(1, 5.25, 0.25)

    S_ds = {}
    for q in tqdm(qs, desc = "Calcolo spettri di spostamento"): # per tutti i valori di q

        # Calcolo dello spettro di spostamento per ogni accelerogramma
        S_ds[f"{q}"] = []
        for id_acc in tqdm(range(1, len(accs_data["acc"]) + 1), leave=False, desc="GM"):
            S_ds[f"{q}"].append(eval_spectrum(id_acc, q, accs_data, Ts=Ts))

        # Calcolo dello spettro medio
        S_ds[f"{q}_avg"] = np.mean(S_ds[f"{q}"], axis=0)

    
        ### Il codice finisce qui, il resto crea i plot e salva i risultati

        ### Plot
        fig = plt.figure(figsize=plt_set.fig_size)
        ax = fig.gca()
        fig, ax = plot_spectrum(fig, ax, S_ds, Ts, q)
        ax.set_title(f'Displ. spectra q={q}')
        plt.tight_layout(), plt.savefig(f'./figs/displ_spectr_q_{q}.png', dpi=250, bbox_inches='tight')
        ax.set_xlim([0.1, 1.5]), ax.set_ylim([0, q*1.75])
        ax.set_title(f'Displ. spectra q={q}, zoom fino a 1.5 s')
        plt.tight_layout(), plt.savefig(f'./figs/displ_spectr_q_{q}_zoom.png', dpi=250, bbox_inches='tight')


    # # Salvataggio del pickle (serve per futuri sviluppi)
    # with open('displ_spectra_results.pkl', 'wb') as f:
    #     pickle.dump(S_ds, f)

    # Esportazione risultati su Excel
    export_to_excel(S_ds, 'displ_spectra_results.xlsx')
